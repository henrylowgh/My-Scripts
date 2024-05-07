import openpyxl
import tkinter as tk
from tkinter import filedialog
from Bio import SeqIO
import re
import os

# Extract sequence names from FASTA and QC Excel
def parse_identifier(full_sequence_name):
    match = re.match(r'^>?(.+)-(H|L)(\d+)', full_sequence_name)
    if match:
        prefix = match.group(1)
        chain_type = match.group(2)
        number = match.group(3)
        return f"{prefix}-{number}", f"{prefix}-{chain_type}{number}", chain_type
    return None, None, None

# Triage sequences to one of 7 categories
'''
			1. CRL >= 500 & QS >= 40
			2. CRL >= 500 & QS of 25-39
			3. CRL >= 500 & QS < 25
			4. CRL < 500 & QS >= 40
			5. CRL < 500 & QS of 25-39
			6. CRL < 500 & QS < 25
            7. Pairs that are missing heavy or light chain sequence OR missing QC data for either sequence 
'''
def determine_category(crl, qs):
    if crl >= 500:
        if qs >= 40:
            return 1
        elif 25 <= qs <= 39:
            return 2
        elif qs < 25:
            return 3
    else:
        if qs >= 40:
            return 4
        elif 25 <= qs <= 39:
            return 5
        elif qs < 25:
            return 6
    return 7  # Default to Category 7 if criteria are not met

# Provide file dialog boxes for users to specify:
# (I) Input directory containing Excel QC files (including files for multiple reads)
# (II) Input directory folder containing FASTA sequence files
# (III) Output directory folder for triaged sequence FASTA files, QC Excel file with category labels, and log.txt   
def process_antibody_data():
    root = tk.Tk()
    root.withdraw()
    QC_file_path = filedialog.askopenfilename(title='Select folder containing Excel QC files', filetypes=[('Excel files', '*.xlsx *.xls')]) # (I)
    fasta_file_path = filedialog.askopenfilename(title='Select folder containing FASTA sequence files', filetypes=[('FASTA files', '*.fasta *.fa')]) # (II)
    output_dir = filedialog.askdirectory(title='Select Output Directory') # (III)
    if not QC_file_path or not fasta_file_path or not output_dir:
        return print("File selection incomplete or incorrect file format, exiting the script.")

    wb = openpyxl.load_workbook(QC_file_path) # Open excel workbook containing QC data
    ws = wb.active # use active worksheet
    header = [cell.value for cell in ws[1]]  # Existing headers from the first row
    if 'Chain Category' not in header:
        header.extend(['Chain Category', 'Pair Category']) # Add new columns for Triage Category labels
        for col_index, header_title in enumerate(header, start=1):
            ws.cell(row=1, column=col_index, value=header_title)

    pairs = {}
    seq_subsets = {i: [] for i in range(1, 8)}  # 8 is not inclusive, therefore this range goes up to Category 7
    debug_output = []

    # Process and read QC data and prepare data structure for light/heavy chain sequence pairing 
    '''
    EXAMPLES
    base_id: TDM-1-1
    full_id: TDM-1-L1
    chain_type: L
    '''
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=False):
        template_name = row[header.index('TemplateName')].value
        crl = row[header.index('CRL')].value
        qs = row[header.index('QualitySCore')].value
        _, full_id, chain_type = parse_identifier(template_name)
        if full_id: # Check if the row has an id
            category = determine_category(crl, qs)
            base_id, _, _ = parse_identifier(template_name) # Parse QC Excel column 'TemplateName'
            if base_id not in pairs: # pairs is a dictionary contain the Category data for different sequence entries
                pairs[base_id] = {'H': 7, 'L': 7}  # Assume missing and assign Category 7 unless found
            '''
            PSEUDOCODE for above line:
            if "TDM-1-2" not in pairs:
                pairs["TDM-1-2"] = {'H': 7, 'L': 7}
            '''
            pairs[base_id][chain_type] = min(pairs[base_id].get(chain_type, 7), category) # Identify highest quality category (smallest #) from multiple reads of a single sequence chain (a.k.a single full_id)
            row[header.index('Chain Category')].value = category  # Add 'Chain Category' value to Excel output file
            debug_output.append(f"TemplateName: {template_name}, CRL: {crl}, QS: {qs}, Chain: {chain_type}, Category: {category}")

    # Set Pair Category once all entries processed and log the final pair category
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=False):
        template_name = row[header.index('TemplateName')].value
        base_id, _, _ = parse_identifier(template_name)
        if base_id in pairs and 'H' in pairs[base_id] and 'L' in pairs[base_id]:
            pair_category = max(pairs[base_id].values()) # Assign lower quality category (larger #) from between the heavy and light chain of the base_id 
            row[header.index('Pair Category')].value = pair_category # Add 'Pair Category' value to Excel output file
            debug_output.append(f"Pair TemplateName: {template_name}, Pair Category: {pair_category}, Determined by: {'H' if pairs[base_id]['H'] == pair_category else 'L'}")

    # Process and read FASTA sequences
    for sequence in SeqIO.parse(fasta_file_path, "fasta"):
        base_id, _, _ = parse_identifier(sequence.id) # Parse FASTA sequence id strings
        if base_id in pairs and 'H' in pairs[base_id] and 'L' in pairs[base_id]:
            assigned_category = max(pairs[base_id].values()) # Assign lower quality category (larger #) from between the heavy and light chain of the base_id 
            seq_subsets[assigned_category].append((sequence.id, str(sequence.seq))) # Add triaged sequence to category-specific FASTA file
            debug_output.append(f"FASTA ID: {sequence.id}, Assigned Category: {assigned_category}")

    # Save sequences to separate output FASTA files in user-designated output directory
    for index, sequences in seq_subsets.items():
        file_name = f'Category_{index}_paired_sequences.fasta'
        with open(os.path.join(output_dir, file_name), 'w') as file:
            for seq_id, seq in sequences:
                file.write(f'>{seq_id}\n{seq}\n')

    # Save the modified Excel workbook (which includes the 2 new Category columns) in user-designated output directory
    new_excel_file_name = 'Modified_QC_Data_with_pairs.xlsx'
    new_file_path = os.path.join(output_dir, new_excel_file_name)
    wb.save(new_file_path)

    # Save debugging output to log file in user-designated output directory
    log_file_path = os.path.join(output_dir, 'log.txt')
    with open(log_file_path, 'w') as log_file:
        log_file.write('\n'.join(debug_output))

    print("Files and logs have been successfully saved to the selected directory.")

if __name__ == "__main__":
    process_antibody_data()