import openpyxl
import tkinter as tk
from tkinter import filedialog
from Bio import SeqIO

def process_antibody_data():
    root = tk.Tk()
    root.withdraw()  # Hide the main window

    input_file_path = filedialog.askopenfilename(
        title='Select the Excel file', filetypes=[('Excel files', '*.xlsx *.xls')])
    fasta_file_path = filedialog.askopenfilename(
        title='Select the FASTA file', filetypes=[('FASTA files', '*.fasta *.fa')])
    output_dir = filedialog.askdirectory(title='Select Output Directory')

    if not input_file_path or not fasta_file_path or not output_dir:
        return print("File selection incomplete, exiting the script.")

    wb = openpyxl.load_workbook(input_file_path)
    ws = wb.active
    data = list(ws.iter_rows(values_only=True))
    header = data[0]
    data = data[1:]
    crl_index = header.index('CRL')
    qs_index = header.index('QualitySCore')
    template_name_index = header.index('TemplateName')

    seq_dict = {record.description: str(record.seq) for record in SeqIO.parse(fasta_file_path, "fasta")}

    subsets = {i: [] for i in range(6)}
    seq_subsets = {i: {} for i in range(6)}  # Use a dictionary to avoid duplicates

    for row in data:
        template_name = row[template_name_index]
        crl = row[crl_index]
        qs = row[qs_index]
        category = determine_category(crl, qs)
        subsets[category].append(row)
        for desc, seq in seq_dict.items():
            if template_name in desc:
                seq_subsets[category][desc] = seq  # This prevents duplicates

    save_subsets(subsets, header, output_dir)

    for index, seqs in seq_subsets.items():
        file_name = f'/Category_{index + 1}.fasta'
        with open(output_dir + file_name, 'w') as file:
            for seq_description, sequence in seqs.items():
                file.write(f'>{seq_description}\n{sequence}\n')

    print("Files have been successfully saved to the selected directory.")

def determine_category(crl, qs):
    if crl >= 500:
        if qs >= 40:
            return 0
        elif 25 <= qs <= 39:
            return 1
        elif qs < 25:
            return 2
    else:
        if qs >= 40:
            return 3
        elif 25 <= qs <= 39:
            return 4
        elif qs < 25:
            return 5

def save_subsets(subsets, header, output_dir):
    file_names = [
        '/Category_I_CRL500_QS40.xlsx',
        '/Category_II_CRL500_QS25-39.xlsx',
        '/Category_III_CRL500_QSlt25.xlsx',
        '/Category_IV_CRLlt500_QS40.xlsx',
        '/Category_V_CRLlt500_QS25-39.xlsx',
        '/Category_VI_CRLlt500_QSlt25.xlsx'
    ]
    for index, rows in subsets.items():
        new_wb = openpyxl.Workbook()
        new_ws = new_wb.active
        new_ws.append(header)
        for row in rows:
            new_ws.append(row)
        new_wb.save(output_dir + file_names[index])

if __name__ == "__main__":
    process_antibody_data()
