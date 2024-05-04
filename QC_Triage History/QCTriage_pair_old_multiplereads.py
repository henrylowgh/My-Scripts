import openpyxl
import tkinter as tk
from tkinter import filedialog
from Bio import SeqIO
import re
import os

def parse_fasta_names(full_sequence_name):
    # Regex to capture identifier and chain type
    match = re.match(r'^>?(.+)-(H|L)(\d+)', full_sequence_name)
    if match:
        prefix = match.group(1)
        chain_type = match.group(2)
        number = match.group(3)
        print(f"FASTA | {prefix}-{chain_type}{number} | Chain: {chain_type}")
        return f"{prefix}-{chain_type}{number}", chain_type
    return None, None

def parse_QC_template_names(template_name):
    # Similar regex as used for FASTA to ensure consistent identifier extraction
    match = re.match(r'^>?(.+)-(H|L)(\d+)', template_name)
    if match:
        prefix = match.group(1)
        chain_type = match.group(2)
        number = match.group(3)
        print(f"QC | {prefix}-{chain_type}{number} | Chain: {chain_type}")
        return f"{prefix}-{chain_type}{number}", chain_type
    return None, None

def determine_category(crl, qs):
    # Category determination logic
    if crl >= 500:
        if qs >= 40:
            return 1
        elif 25 <= qs <= 39:
            return 2
        elif qs < 25:
            return 3
    else:
        if qs >= 40:
            return 4
        elif 25 <= qs <= 39:
            return 5
        elif qs < 25:
            return 6

def process_antibody_data():
    root = tk.Tk()
    root.withdraw()

    QC_file_path = filedialog.askopenfilename(title='Select the Excel file', filetypes=[('Excel files', '*.xlsx *.xls')])
    fasta_file_path = filedialog.askopenfilename(title='Select the FASTA file', filetypes=[('FASTA files', '*.fasta *.fa')])
    output_dir = filedialog.askdirectory(title='Select Output Directory')

    if not QC_file_path or not fasta_file_path or not output_dir:
        return print("File selection incomplete, exiting the script.")

    wb = openpyxl.load_workbook(QC_file_path)
    ws = wb.active
    header = [cell.value for cell in ws[1]]
    data = [list(row) for row in ws.iter_rows(values_only=True, min_row=2)]

    crl_index = header.index('CRL')
    qs_index = header.index('QualitySCore')
    sequence_name_index = header.index('TemplateName')

    header.append('Category')
    seq_subsets = {i: [] for i in range(1, 7)}
    lowest_category = {}

    for row in data:
        template_name = row[sequence_name_index]
        extracted_template_name, QC_chain_type = parse_QC_template_names(template_name)
        if extracted_template_name:
            crl = row[crl_index]
            qs = row[qs_index]
            category = determine_category(crl, qs)
            if extracted_template_name in lowest_category:
                if category < lowest_category[extracted_template_name]:
                    lowest_category[extracted_template_name] = category
                    print(f"Updated {extracted_template_name} to a lower category {category}")
            else:
                lowest_category[extracted_template_name] = category
                print(f"Assigned {extracted_template_name} to category {category}")
            row.append(lowest_category[extracted_template_name])

    for col_num, header_title in enumerate(header, start=1):
        ws.cell(row=1, column=col_num, value=header_title)
    for i, row in enumerate(data, start=2):
        for j, value in enumerate(row):
            ws.cell(row=i, column=j+1, value=value)

    new_excel_file_name = 'Modified_QC_Data.xlsx'
    new_file_path = os.path.join(output_dir, new_excel_file_name)
    wb.save(new_file_path)

    for sequence in SeqIO.parse(fasta_file_path, "fasta"):
        id_key, FASTA_chain_type = parse_fasta_names(sequence.id)
        if id_key and id_key in lowest_category:
            sequence_id, seq = sequence.id, str(sequence.seq)
            assigned_category = lowest_category[id_key]
            seq_subsets[assigned_category].append((sequence_id, seq))
            print(f"Matched {sequence_id} with Excel entry in lowest Category {assigned_category}")

    for index, sequences in seq_subsets.items():
        file_name = f'Category_{index}_sequences.fasta'
        with open(os.path.join(output_dir, file_name), 'w') as file:
            for id, seq in sequences:
                file.write(f'>{id}\n{seq}\n')

    print("Files have been successfully saved to the selected directory.")

if __name__ == "__main__":
    process_antibody_data()
