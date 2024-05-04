import openpyxl
import tkinter as tk
from tkinter import filedialog
from Bio import SeqIO
import re
import os

def parse_identifier(full_sequence_name):
    match = re.match(r'^>?(.+)-(H|L)(\d+)', full_sequence_name)
    if match:
        prefix = match.group(1)
        chain_type = match.group(2)
        number = match.group(3)
        return f"{prefix}-{number}", f"{prefix}-{chain_type}{number}", chain_type
    return None, None, None

def determine_category(crl, qs):
    if crl >= 500:
        if qs >= 40:
            return 1
        elif 25 <= qs <= 39:
            return 2
        elif qs < 25:
            return 3
    else:
        if qs >= 40:
            return 4
        elif 25 <= qs <= 39:
            return 5
        elif qs < 25:
            return 6

def process_antibody_data():
    root = tk.Tk()
    root.withdraw()
    QC_file_path = filedialog.askopenfilename(title='Select the Excel file', filetypes=[('Excel files', '*.xlsx *.xls')])
    fasta_file_path = filedialog.askopenfilename(title='Select the FASTA file', filetypes=[('FASTA files', '*.fasta *.fa')])
    output_dir = filedialog.askdirectory(title='Select Output Directory')
    if not QC_file_path or not fasta_file_path or not output_dir:
        return print("File selection incomplete, exiting the script.")

    wb = openpyxl.load_workbook(QC_file_path)
    ws = wb.active
    header = [cell.value for cell in ws[1]]  # Existing headers from the first row
    # Check if new headers are already added
    if 'Chain Category' not in header:
        header.extend(['Chain Category', 'Pair Category'])  # Extend with new column titles
        # Update the header row
        for col_index, header_title in enumerate(header, start=1):
            ws.cell(row=1, column=col_index, value=header_title)

    pairs = {}
    lowest_category = {}

    # Processing each row and appending data
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=False):
        template_name = row[header.index('TemplateName')].value
        crl = row[header.index('CRL')].value
        qs = row[header.index('QualitySCore')].value
        _, identifier, chain_type = parse_identifier(template_name)
        if identifier:
            category = determine_category(crl, qs)
            base_id, full_id = parse_identifier(template_name)[:2]
            if base_id not in pairs:
                pairs[base_id] = {}
            pairs[base_id][chain_type] = min(pairs[base_id].get(chain_type, float('inf')), category)
            lowest_category[full_id] = pairs[base_id][chain_type]
            # Set values directly under the new headers
            row[len(header) - 2].value = category  # Chain Category
            row[len(header) - 1].value = max(pairs[base_id].values())  # Pair Category

    seq_subsets = {i: [] for i in range(1, 7)}
    for sequence in SeqIO.parse(fasta_file_path, "fasta"):
        base_id, full_id, chain_type = parse_identifier(sequence.id)
        if base_id and full_id in lowest_category:
            assigned_category = max(pairs[base_id].values())
            seq_subsets[assigned_category].append((sequence.id, str(sequence.seq)))

    # Save sequences
    for index, sequences in seq_subsets.items():
        file_name = f'Category_{index}_paired_sequences.fasta'
        with open(os.path.join(output_dir, file_name), 'w') as file:
            for seq_id, seq in sequences:
                file.write(f'>{seq_id}\n{seq}\n')

    # Save the modified Excel workbook
    new_excel_file_name = 'Modified_QC_Data_with_pairs.xlsx'
    new_file_path = os.path.join(output_dir, new_excel_file_name)
    wb.save(new_file_path)
    print("Files have been successfully saved to the selected directory.")

if __name__ == "__main__":
    process_antibody_data()
