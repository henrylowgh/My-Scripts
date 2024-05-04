import openpyxl
import tkinter as tk
from tkinter import filedialog
from Bio import SeqIO
import re
import os

# Search for the regular expression (a.k.a. partial string subset) that matches '[PREFIX]-L1/[PREFIX]-H1' formatting in the FASTA sequence files
def parse_fasta_names(full_sequence_name):
    # Regex updated to optionally ignore a leading '>' and capture the rest of the string up to the digit immediately after the H or L
    match = re.match(r'^>?(.+)-(H|L)(\d+)', full_sequence_name)
    if match:
        prefix = match.group(1)  # This [PREFIX] will include everything up to the hyphen before "H" or "L"
        chain_type = match.group(2) # This will be either H (for heavy chain) or L (for light chain)
        number = match.group(3) # This will be the specific number of the sequence pair
        print(f"FASTA | {prefix}-{chain_type}{number} | Chain: {chain_type}") # Debugging
        return f"{prefix}-{chain_type}{number}", chain_type  # Return both full key and chain type
    return None, None  # Return None for both if no regular expression match

# Function to extract the TemplateName column from the Excel file for matching to the extracted FASTA sequence names
def parse_QC_template_names(template_name):
    match = re.match(r'^>?(.+)-(H|L)(\d+)', template_name)
    if match:
        prefix = match.group(1)
        chain_type = match.group(2)
        number = match.group(3)
        print(f"QC | {prefix}-{chain_type}{number} | Chain: {chain_type}") # Debugging
        return f"{prefix}-{chain_type}{number}", chain_type
    return None, None 

# Assign into different categories based on relevant QC parameters
def determine_category(crl, qs):
    if crl >= 500:
        if qs >= 40:
            return 1
        elif 25 <= qs <= 39:
            return 2
        elif qs < 25:
            return 3
    else:
        if qs >= 40:
            return 4
        elif 25 <= qs <= 39:
            return 5
        elif qs < 25:
            return 6

def process_antibody_data():
    root = tk.Tk()
    root.withdraw()

    # Dialog box options
    QC_file_path = filedialog.askopenfilename(
        title='Select the Excel file', filetypes=[('Excel files', '*.xlsx *.xls')])
    fasta_file_path = filedialog.askopenfilename(
        title='Select the FASTA file', filetypes=[('FASTA files', '*.fasta *.fa')])
    output_dir = filedialog.askdirectory(title='Select Output Directory')

    if not QC_file_path or not fasta_file_path or not output_dir:
        return print("File selection incomplete, exiting the script.")

    wb = openpyxl.load_workbook(QC_file_path)
    ws = wb.active
    header = [cell.value for cell in ws[1]]  # Ensure header is a list
    data = [list(row) for row in ws.iter_rows(values_only=True, min_row=2)]  # Convert all data rows to lists
    
    crl_index = header.index('CRL')
    qs_index = header.index('QualitySCore')
    sequence_name_index = header.index('TemplateName')

    # Adding a new column for category to the existing data
    header.append('Category')
    chains = {}
    seq_subsets = {i: [] for i in range(1, 7)}  # 6 categories (1-6)
    qc_subsets = {i: [] for i in range(1, 7)}  # 6 categories (1-6)


    '''Analyze Excel QC data'''
    # Analyze Excel QC data entries
    for row in data:
        template_name = row[sequence_name_index]
        extracted_template_name, QC_chain_type = parse_QC_template_names(template_name)
        if extracted_template_name:
            crl = row[crl_index]
            qs = row[qs_index]
            category = determine_category(crl, qs)
            chains[extracted_template_name] = (template_name, category)
            row.append(category)  # Append category directly to each row
    
    # Writing QC data to new combined Excel file (with category labels)
    for col_num, header_title in enumerate(header, start=1): # Write the updated header to the modified workbook
        ws.cell(row=1, column=col_num, value=header_title)
    
    for i, row in enumerate(data, start=2): # Write the updated data to the modified workbook
        for j, value in enumerate(row):
            ws.cell(row=i, column=j+1, value=value) # Update cells in the workbook with the modified data

    new_excel_file_name = 'Modified_QC_Data.xlsx' # Define a new file name for the modified Excel workbook
    # Combine the new file name with the output directory to create the full path
    new_file_path = os.path.join(output_dir, new_excel_file_name)
    # Save the modified workbook to the new file path in the output directory
    wb.save(new_file_path)  # Save the workbook to the new file path

    # Save QC entries in separate triaged Excel files
    for index, qc_data in qc_subsets.items():
        file_name = f'/Category_{index}_QC.xlsx'
        new_wb = openpyxl.Workbook()
        new_ws = new_wb.active
        new_ws.append(header)
        for row in qc_data:
            new_ws.append(row)
        new_wb.save(output_dir + file_name)


    '''Analyze FASTA sequences'''
    for sequence in SeqIO.parse(fasta_file_path, "fasta"): # uses BioPython SeqIO FASTA parsing tool
        id_key, FASTA_chain_type = parse_fasta_names(sequence.id)
        if id_key and id_key in chains:
            sequence_id, seq = sequence.id, str(sequence.seq)
            template_name, assigned_category = chains[id_key]
            seq_subsets[assigned_category].append((sequence_id, seq))
            qc_subsets[assigned_category].extend([row for row in data if row[sequence_name_index] == template_name])
            print(f"Matched {sequence_id} with Excel entry {template_name} in Category {assigned_category}") # Debugging

    # Save FASTA sequences in separate triaged FASTA files
    for index, sequences in seq_subsets.items():
        file_name = f'/Category_{index}_sequences.fasta'
        with open(output_dir + file_name, 'w') as file:
            for id, seq in sequences:
                file.write(f'>{id}\n{seq}\n')

    print("Files have been successfully saved to the selected directory.")


# Run script
if __name__ == "__main__":
    process_antibody_data()
    
