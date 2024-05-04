import openpyxl
import tkinter as tk
from tkinter import filedialog
from Bio import SeqIO
import re
import os

def parse_identifier(full_sequence_name):
    match = re.match(r'^>?(.+)-(H|L)(\d+)', full_sequence_name)
    if match:
        prefix = match.group(1)
        chain_type = match.group(2)
        number = match.group(3)
        return f"{prefix}-{number}", f"{prefix}-{chain_type}{number}", chain_type
    return None, None, None

def determine_category(crl, qs):
    if crl >= 500:
        if qs >= 40:
            return 1
        elif 25 <= qs <= 39:
            return 2
        elif qs < 25:
            return 3
    else:
        if qs >= 40:
            return 4
        elif 25 <= qs <= 39:
            return 5
        elif qs < 25:
            return 6
    return 7  # Default to Category 7 if criteria are not met

def process_antibody_data():
    root = tk.Tk()
    root.withdraw()
    QC_file_path = filedialog.askopenfilename(title='Select the Excel file', filetypes=[('Excel files', '*.xlsx *.xls')])
    fasta_file_path = filedialog.askopenfilename(title='Select the FASTA file', filetypes=[('FASTA files', '*.fasta *.fa')])
    output_dir = filedialog.askdirectory(title='Select Output Directory')
    if not QC_file_path or not fasta_file_path or not output_dir:
        return print("File selection incomplete, exiting the script.")

    wb = openpyxl.load_workbook(QC_file_path)
    ws = wb.active
    header = [cell.value for cell in ws[1]]  # Existing headers from the first row
    if 'Chain Category' not in header:
        header.extend(['Chain Category', 'Pair Category'])
        for col_index, header_title in enumerate(header, start=1):
            ws.cell(row=1, column=col_index, value=header_title)

    pairs = {}
    seq_subsets = {i: [] for i in range(1, 8)}  # Include Category 7

    # Process QC data and prepare structure
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=False):
        template_name = row[header.index('TemplateName')].value
        crl = row[header.index('CRL')].value
        qs = row[header.index('QualitySCore')].value
        _, identifier, chain_type = parse_identifier(template_name)
        if identifier:
            category = determine_category(crl, qs)
            base_id, full_id = parse_identifier(template_name)[:2]
            if base_id not in pairs:
                pairs[base_id] = {'H': 7, 'L': 7}  # Assume missing unless found
            pairs[base_id][chain_type] = min(pairs[base_id].get(chain_type, 7), category)
            row[header.index('Chain Category')].value = category  # Update chain category

    # Set Pair Category once all entries processed
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=False):
        template_name = row[header.index('TemplateName')].value
        base_id, _ = parse_identifier(template_name)[:2]
        if base_id and base_id in pairs:
            if 'H' in pairs[base_id] and 'L' in pairs[base_id]:
                pair_category = max(pairs[base_id].values())
                row[header.index('Pair Category')].value = pair_category

    # Process FASTA sequences
    for sequence in SeqIO.parse(fasta_file_path, "fasta"):
        base_id, full_id, chain_type = parse_identifier(sequence.id)
        if base_id in pairs and 'H' in pairs[base_id] and 'L' in pairs[base_id]:
            assigned_category = max(pairs[base_id].values())
            seq_subsets[assigned_category].append((sequence.id, str(sequence.seq)))
        else:
            seq_subsets[7].append((sequence.id, str(sequence.seq)))  # Handle missing cases

    # Save sequences
    for index, sequences in seq_subsets.items():
        file_name = f'Category_{index}_paired_sequences.fasta'
        with open(os.path.join(output_dir, file_name), 'w') as file:
            for seq_id, seq in sequences:
                file.write(f'>{seq_id}\n{seq}\n')

    # Save the modified Excel workbook
    new_excel_file_name = 'Modified_QC_Data_with_pairs.xlsx'
    new_file_path = os.path.join(output_dir, new_excel_file_name)
    wb.save(new_file_path)
    print("Files have been successfully saved to the selected directory.")

if __name__ == "__main__":
    process_antibody_data()
