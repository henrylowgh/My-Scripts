import openpyxl
import tkinter as tk
from tkinter import filedialog
from Bio import SeqIO
import re
import os
import glob
import pandas as pd
import matplotlib.pyplot as plt # type: ignore

# Extract sequence names from FASTA and QC Excel
def parse_identifier(full_sequence_name):
    match = re.match(r'^>?([\w-]+?)-?(b|a)(\d+)', full_sequence_name)
    if match:
        prefix = match.group(1)
        chain_type = match.group(2)
        number = match.group(3)
        return f"{prefix}-{number}", f"{prefix}-{number}{chain_type}", chain_type
    return None, None, None

# Triage sequences to one of 7 categories
'''
			1. CRL >= 500 & QS >= 40
			2. CRL >= 500 & QS of 25-39
			3. CRL >= 500 & QS < 25
			4. CRL < 500 & QS >= 40
			5. CRL < 500 & QS of 25-39
			6. CRL < 500 & QS < 25
            7. Pairs that are missing heavy or light chain FASTA sequence OR missing QC data for either sequence 
'''
def determine_category(crl, qs):
    # Ensure that CRL and QualityScore are not None and are integers
    if crl is None or qs is None:
        return 7  # Default to Category 7 if CRL or QS data is missing or invalid
    
    try:
        crl = int(crl)
        qs = int(qs)
    except ValueError:
        return 7  # Handle cases where CRL or QS cannot be converted to integers

    if crl >= 500:
        if qs >= 40:
            return 1
        elif 25 <= qs <= 39:
            return 2
        elif qs < 25:
            return 3
    else:
        if qs >= 40:
            return 4
        elif 25 <= qs <= 39:
            return 5
        elif qs < 25:
            return 6
    return 7  # Default to Category 7 if above criteria are not met

def convert_xls_to_xlsx(xls_path):
# Define the new .xlsx file path
    xlsx_path = xls_path + 'x'
    # Read the xls file using pandas
    df = pd.read_excel(xls_path)
    # Save it as xlsx using the openpyxl engine
    df.to_excel(xlsx_path, index=False, engine='openpyxl')
    # Remove the original .xls file
    os.remove(xls_path)
    return xlsx_path  # Return the path to the new .xlsx file

def save_histogram(category_counts, output_dir):
    categories = list(category_counts.keys())
    counts = [category_counts[cat] for cat in categories]

    plt.figure(figsize=(10, 6))
    plt.bar(categories, counts, color='skyblue')
    plt.xlabel('Category')
    plt.ylabel('Number of Pairs')
    plt.title('Distribution of Sequence Pairs by Category')
    plt.xticks(categories)
    plt.grid(axis='y', linestyle='--', alpha=0.7)

    histogram_path = os.path.join(output_dir, 'category_distribution_histogram.png')
    plt.savefig(histogram_path)
    print(f"Histogram saved to {histogram_path}")
    plt.close()

def print_final_pair_categories(pairs, debug_stream):
    print("Final Categories for Sequence Pairs:")
    for base_id, categories in pairs.items():
        if 'b' in categories and 'a' in categories:
            final_category = max(categories.values())
            print(f"Pair {base_id} (Heavy Chain: Category {categories['b']}, Light Chain: Category {categories['a']}) - Final Category: {final_category}")
            debug_stream.append(f"Pair {base_id} (Heavy Chain: Category {categories['b']}, Light Chain: Category {categories['a']}) - Final Category: {final_category}")
        elif 'b' in categories:
            print(f"Pair {base_id} (Heavy Chain: Category {categories['b']}, Light Chain: MISSING) - Final Category: 7")
            debug_stream.append(f"Pair {base_id} (Heavy Chain: Category {categories['b']}, Light Chain: MISSING) - Final Category: 7")
        elif 'a' in categories:
            print(f"Pair {base_id} (Heavy Chain: MISSING, Light Chain: Category {categories['a']}) - Final Category: 7")
            debug_stream.append(f"Pair {base_id} (Heavy Chain: MISSING, Light Chain: Category {categories['a']}) - Final Category: 7")
        else:
            print(f"Pair {base_id} (Heavy Chain: MISSING, Light Chain: MISSING) - Final Category: 7")
            debug_stream.append(f"Pair {base_id} (Heavy Chain: MISSING, Light Chain: MISSING) - Final Category: 7")


# Provide file dialog boxes for users to specify:
# (I) Input directory containing Excel QC files (including files for multiple reads)
# (II) Input directory folder containing FASTA sequence files
# (III) Output directory folder for triaged sequence FASTA files, QC Excel file with category labels, and log.txt 
# Note: Script will combine Excel QC files into a single Excel file, and FASTA sequence files into a single FASTA file, exporting both to the output directory   
def process_antibody_data():
    root = tk.Tk()
    root.withdraw()
    QC_file_dir = filedialog.askdirectory(title='Select directory containing Excel QC files') # (I)
    fasta_file_dir = filedialog.askdirectory(title='Select directory containing FASTA sequence files') # (II)
    output_dir = filedialog.askdirectory(title='Select Output Directory') # (III)
    if not QC_file_dir or not fasta_file_dir or not output_dir:
        return print("Directory selection incomplete or incorrect file format, exiting the script.")


    '''COMBINE FILES FROM INPUT DIRECTORIES'''
    # Initialize main workbook and sheet
    main_wb = openpyxl.Workbook()
    main_ws = main_wb.active
    main_ws.title = "Combined QC Data"
    first_file = True

    # convert .xls to .xlsx as needed
    for file_path in glob.glob(os.path.join(QC_file_dir, '*.*')):
        if file_path.endswith('.xls'):
            file_path = convert_xls_to_xlsx(file_path) # Convert .xls to .xlsx, and delete old .xls files

    # Process all .xlsx Excel files
    for file_path in glob.glob(os.path.join(QC_file_dir, '*.*')):
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active
        if first_file:
            header = [cell.value for cell in ws[1]]  # Extract header from the first file
            main_ws.append(header)  # Append the header to the main worksheet
            first_file = False
        for row in ws.iter_rows(min_row=2, values_only=True):  # Skip the header row for remaining files
            main_ws.append(row)

    # Ensure header is added
    if main_ws.max_row == 1:  # Only header row present
        for cell in ws[1]:  # Take header from the last processed workbook
            main_ws.cell(row=1, column=cell.column, value=cell.value)

    # Process all FASTA files
    all_sequences = []
    file_extensions = ['*.fasta', '*.txt']
    for file_pattern in file_extensions:
        for fasta_file in glob.glob(os.path.join(fasta_file_dir, file_pattern)):
            # print(f"Reading file: {fasta_file}")  # Debug print to check if files are being read
            for record in SeqIO.parse(fasta_file, "fasta"):
                all_sequences.append(record)
            print(f"Found {len(all_sequences)} sequences after reading {fasta_file}")  # Debug print to check sequence accumulation

        
    # Save the combined Excel workbook
    combined_excel_path = os.path.join(output_dir, "Combined_qc_data.xlsx")
    main_wb.save(combined_excel_path)

    # Save all sequences to a single FASTA file
    combined_fasta_path = os.path.join(output_dir, "Combined_sequences.fasta")
    SeqIO.write(all_sequences, combined_fasta_path, "fasta")


    '''BEGIN PROCESSING COMBINED INPUT FILES'''
    wb = openpyxl.load_workbook(combined_excel_path) # Open excel workbook containing QC data
    ws = wb.active # use active worksheet
    header = [cell.value for cell in ws[1]]  # Existing headers from the first row
    if 'Chain Category' not in header:
        header.extend(['Chain Category', 'Pair Category']) # Add new columns for Triage Category labels
        for col_index, header_title in enumerate(header, start=1):
            ws.cell(row=1, column=col_index, value=header_title)

    pairs = {} # Initialize pairs dictionary, which contains a category value for the heavy and light chain for each sequence id key
    seq_subsets = {i: [] for i in range(1, 8)}  # 8 is not inclusive, therefore this range goes up to Category 7
    debug_output = [] # Initialize debugging output list stream
    fasta_sequence_ids = set()  # Set to track sequence IDs from FASTA files

    '''
    VARIABLE EXAMPLES
    base_id: B1-1 (prefix-number)
    full_id: B1-1L (prefix-number_chain type)
    chain_type: L (chain type)

    base_id: B1-1 (prefix-number)
    full_id: B1-1a (prefix-number_chain type)
    chain_type: a (chain type)

    a = Light chain
    b = Heavy chain
    '''

    # Process QC data entries and assign Chain Category in the QC data (but not Pair Category yet)
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=False):
        template_name = row[header.index('TemplateName')].value
        crl = row[header.index('CRL')].value
        qs = row[header.index('QualitySCore')].value
        base_id, full_id, chain_type = parse_identifier(template_name)
        if full_id: # Check if the row has an id
            category = determine_category(crl, qs)
            if base_id not in pairs: 
                pairs[base_id] = {'b': 7, 'a': 7}  # Initialize as Category 7 if TemplateName id not yet an entry (applies to QC entries only, at this point)
            pairs[base_id][chain_type] = min(pairs[base_id].get(chain_type, 7), category) # Set as highest quality category (smallest #) from multiple reads of a single sequence chain (a.k.a single full_id)
            row[header.index('Chain Category')].value = category  # Add 'Chain Category' value to Excel output file
            debug_output.append(f"QC Entry: {template_name}, CRL: {crl}, QS: {qs}, Chain: {chain_type}, Chain Category: {category}")

    # Identify which FASTA sequences are missing QC entries (don't initialize yet since that throws off the debugging output)
    for sequence in SeqIO.parse(combined_fasta_path, "fasta"):
        base_id, full_id, chain_type = parse_identifier(sequence.id) # Parse FASTA sequence id strings   
        fasta_sequence_ids.add(full_id) # keep track of full_ID (i.e. specific chains)
        # print(f"debug: {full_id}")
        if base_id not in pairs: # a.k.a sequences that are missing QC entries
            print(f"{base_id} pair missing QC entry!")
            debug_output.append(f"{base_id} pair missing QC entry!") 

    # print(fasta_sequence_ids)
    # Identify which QC entries are missing FASTA sequences
    for base_id in pairs:
        for chain_type in ['b', 'a']:
            full_id = f"{base_id}{chain_type}"
            # print(full_id)
            if full_id not in fasta_sequence_ids:
                pairs[base_id][chain_type] = 7
                print(f"{chain_type}{base_id} QC entry has no matching FASTA sequence!")
                debug_output.append(f"{base_id}{chain_type} QC entry has no matching FASTA sequence!")

    # Identify Pair Category for QC Pairs
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=False):
        template_name = row[header.index('TemplateName')].value
        base_id, _, _ = parse_identifier(template_name)
        if base_id in pairs and 'b' in pairs[base_id] and 'a' in pairs[base_id]:
            pair_category = max(pairs[base_id].values()) # Assign lower quality category (larger #) from between the heavy and light chain of the base_id 
            row[header.index('Pair Category')].value = pair_category # Add 'Pair Category' value to Excel output file
            debug_output.append(f"QC H/L Chain Pairing: {template_name}, Pair Category: {pair_category}, Determined by: {'H' if pairs[base_id]['b'] == pair_category else 'L'}")

    # Identify Pair Category for FASTA Pairs
    for sequence in SeqIO.parse(combined_fasta_path, "fasta"):
        base_id, full_id, chain_type = parse_identifier(sequence.id) # Parse FASTA sequence id strings      
        if base_id not in pairs: 
            pairs[base_id] = {'b': 7, 'a': 7}  # At this point, initialize missing pairs if not in pairs from Excel , these will be same as "pair missing QC entry"
        pair_category = max(pairs[base_id].values()) # Assign lower quality category (larger #) from between the heavy and light chain of the base_id 
        seq_subsets[pair_category].append((sequence.id, str(sequence.seq))) # Add triaged sequence to category-specific FASTA file
        debug_output.append(f"FASTA Sequence: {sequence.id}, Pair Category: {pair_category}")
    
    # Save sequences to separate output FASTA files in user-designated output directory
    for index, sequences in seq_subsets.items():
        file_name = f'Category_{index}_paired_sequences.fasta'
        with open(os.path.join(output_dir, file_name), 'w') as file:
            for seq_id, seq in sequences:
                file.write(f'>{seq_id}\n{seq}\n')

    category_counts = {i: 0 for i in range(1, 8)}  # Initialize counts to 0 for categories 1 to 7

    # Count up number of pairs in each category
    for base_id, categories in pairs.items():
        if 'b' in categories and 'a' in categories:
            final_category = max(categories.values())
            category_counts[final_category] += 1
    # Print category statistics and export statistics to log
    total_pairs = sum(category_counts.values())
    print("Category Statistics:")
    debug_output.append("Category Statistics:")
    for category, count in category_counts.items():
        percent = (count / total_pairs * 100) if total_pairs > 0 else 0
        print(f"Category {category}: {count} pairs, {percent:.2f}%")
        debug_output.append(f"Category {category}: {count} pairs, {percent:.2f}%")
    print(f"Total Pairs: {total_pairs}")
    debug_output.append(f"Total Pairs: {total_pairs}")

    # Save the modified Excel workbook (which includes the 2 new Category columns) in user-designated output directory
    new_excel_file_name = 'COMBINED_QC_DATA_WITH_CATEGORIES.xlsx'
    new_file_path = os.path.join(output_dir, new_excel_file_name)
    wb.save(new_file_path)    

    # Save debugging output to log file in user-designated output directory
    log_file_path = os.path.join(output_dir, 'Triage_log.txt')
    with open(log_file_path, 'w') as log_file:
        log_file.write('\n'.join(debug_output))
    save_histogram(category_counts, output_dir) # generate and save histogram of results to output directory

    print("Files and logs have been successfully saved to the selected directory.")

    # DEBUGGING
    # print_final_pair_categories(pairs, debug_output)
    # print(f"Pairs: {pairs}")

if __name__ == "__main__":
    process_antibody_data()